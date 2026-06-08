"""
import_rh.py

Script untuk import data RH dari file Excel BPS
ke tabel rh_komoditas di database.

Cara pakai:
    python import_rh.py

Letakkan file ini satu folder dengan app.py.
Pastikan file Excel ada di path yang ditentukan di EXCEL_PATH.

v1.0
"""

import os
import sys
import openpyxl
import pandas as pd
from datetime import date
from dotenv import load_dotenv
from sqlalchemy import create_engine, text

load_dotenv()

# ═══════════════════════════════════════════════════════════════
# KONFIGURASI — Sesuaikan path file Excel dan sheet yang dipakai
# ═══════════════════════════════════════════════════════════════

# Path ke file Form Input RH dari BPS
EXCEL_PATH = "Form Input RH.xlsx"

# Sheet yang dipakai — sesuaikan dengan kota/wilayahmu
# Pilihan: '1403' (Tembilahan), '1406', '1471', '1473'
# Untuk Riau gabungan, pilih sheet yang paling lengkap
SHEET_NAME = "1471"

# Mapping nama komoditas di Excel -> komoditas_id di database
# Kamu perlu isi ini sesuai data master_komoditas-mu
# Format: 'NAMA DI EXCEL': komoditas_id
KOMODITAS_MAP = {
    # Isi sesuai nama_komoditas di master_komoditas
    # Contoh (sesuaikan dengan data aktualmu):
    'BERAS':                                          1,
    'TEPUNG TERIGU':                                  2,
    'DAGING AYAM RAS':                                3,
    'DAGING SAPI':                                    4,
    'IKAN KEMBUNG/IKAN GEMBUNG/ IKAN BANYAR/IKAN GEMBOLO/ IKAN ASO-ASO': 5,
    'SUSU BUBUK':                                     6,
    'SUSU KENTAL MANIS':                              7,
    'TELUR AYAM RAS':                                 8,
    'MINYAK GORENG':                                  9,
    'TOMAT':                                          10,
    'CABAI MERAH':                                    11,
    'CABAI RAWIT':                                    12,
    'BAWANG MERAH':                                   13,
    'BAWANG PUTIH':                                   14,
    'GULA PASIR':                                     15,
    'TEPUNG TERIGU':                                  16,
    'BAHAN BAKAR RUMAH TANGGA':                       20,
    'SUSU BUBUK UNTUK BAYI':                          19,
    # Tambahkan mapping lainnya sesuai kebutuhan
}


# ═══════════════════════════════════════════════════════════════
# DATABASE CONNECTION
# ═══════════════════════════════════════════════════════════════

def get_engine():
    db_url = os.getenv('DATABASE_URL')
    if not db_url:
        host     = os.getenv('DB_HOST', 'localhost')
        port     = os.getenv('DB_PORT', '3306')
        user     = os.getenv('DB_USER', 'root')
        password = os.getenv('DB_PASSWORD', '')
        dbname   = os.getenv('DB_NAME', 'commodityapp')
        db_url   = f"mysql+pymysql://{user}:{password}@{host}:{port}/{dbname}"
    return create_engine(db_url)


# ═══════════════════════════════════════════════════════════════
# HELPER — Parse periode dari header kolom
# ═══════════════════════════════════════════════════════════════

def parse_periode(header: str):
    """
    Parse 'RH 2301' -> date(2023, 1, 1)
    Parse 'RH 2412' -> date(2024, 12, 1)
    """
    if not header or not str(header).startswith('RH '):
        return None
    try:
        kode = str(header).replace('RH ', '').strip()
        if len(kode) != 4:
            return None
        tahun = int('20' + kode[:2])
        bulan = int(kode[2:])
        if not (1 <= bulan <= 12):
            return None
        return date(tahun, bulan, 1)
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════════
# STEP 1 — Baca mapping komoditas dari database
# ═══════════════════════════════════════════════════════════════

def load_komoditas_from_db(engine) -> dict:
    """
    Ambil semua komoditas dari DB dan buat mapping
    nama_komoditas (uppercase) -> komoditas_id
    """
    query = "SELECT id, nama_komoditas FROM master_komoditas"
    with engine.connect() as conn:
        rows = conn.execute(text(query)).fetchall()

    mapping = {}
    for row in rows:
        nama = str(row[1]).strip().upper()
        mapping[nama] = int(row[0])

    print(f"[DB] {len(mapping)} komoditas dimuat dari database:")
    for nama, cid in mapping.items():
        print(f"     ID {cid:2d} -> {nama}")
    return mapping


# ═══════════════════════════════════════════════════════════════
# STEP 2 — Baca data RH dari Excel
# ═══════════════════════════════════════════════════════════════

def read_rh_from_excel(excel_path: str, sheet_name: str) -> pd.DataFrame:
    """
    Baca data RH dari file Excel BPS.

    Returns:
        DataFrame dengan kolom: nama_komoditas, tanggal, nilai_rh
    """
    print(f"\n[Excel] Membaca {excel_path} sheet '{sheet_name}'...")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' tidak ditemukan. "
                         f"Sheet yang tersedia: {wb.sheetnames}")

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))

    # Cari baris header (yang mengandung 'RH 2201' dst)
    # Harus ada minimal 3 kolom yang dimulai 'RH ' + 4 digit angka
    header_row_idx = None
    for i, row in enumerate(all_rows):
        if not row:
            continue
        rh_cols = [c for c in row if c and str(c).startswith('RH ') and len(str(c)) == 7 and str(c)[3:].isdigit()]
        if len(rh_cols) >= 3:
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("Baris header 'RH YYMM' tidak ditemukan di sheet ini.")

    headers = all_rows[header_row_idx]
    print(f"[Excel] Header ditemukan di baris {header_row_idx + 1}")

    # Parse kolom periode
    col_periodes = {}  # col_index -> date
    for col_idx, h in enumerate(headers):
        if h and str(h).startswith('RH '):
            tgl = parse_periode(str(h))
            if tgl:
                col_periodes[col_idx] = tgl

    print(f"[Excel] {len(col_periodes)} periode ditemukan: "
          f"{min(col_periodes.values())} s/d {max(col_periodes.values())}")

    # Cari kolom nama komoditas (kolom 'Komoditas')
    nama_col_idx = None
    for col_idx, h in enumerate(headers):
        if h and str(h).strip().lower() == 'komoditas':
            nama_col_idx = col_idx
            break

    if nama_col_idx is None:
        # Fallback: kolom ke-2 (index 2) biasanya nama komoditas
        nama_col_idx = 2
        print(f"[Excel] Kolom 'Komoditas' tidak ditemukan, pakai kolom index {nama_col_idx}")

    # Baca data baris per baris
    records = []
    data_rows = all_rows[header_row_idx + 1:]

    for row in data_rows:
        if not row or row[nama_col_idx] is None:
            continue

        nama = str(row[nama_col_idx]).strip().upper()
        if not nama or nama == 'NONE':
            continue

        for col_idx, tgl in col_periodes.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None:
                continue
            try:
                rh = float(val)
                if rh <= 0:
                    continue
                records.append({
                    'nama_komoditas': nama,
                    'tanggal':        tgl,
                    'nilai_rh':       rh,
                })
            except (TypeError, ValueError):
                continue

    df = pd.DataFrame(records)
    print(f"[Excel] {len(df)} records RH dibaca dari Excel")
    print(f"[Excel] Komoditas unik: {df['nama_komoditas'].nunique()}")
    return df


# ═══════════════════════════════════════════════════════════════
# STEP 3 — Matching nama komoditas Excel -> ID database
# ═══════════════════════════════════════════════════════════════

def match_komoditas(df: pd.DataFrame, db_mapping: dict) -> pd.DataFrame:
    """
    Cocokkan nama komoditas dari Excel dengan ID di database.
    Pakai exact match dulu, kalau tidak ada pakai partial match.
    """
    def find_id(nama_excel):
        # Exact match
        if nama_excel in db_mapping:
            return db_mapping[nama_excel]
        # Partial match — cari nama DB yang ada di nama Excel atau sebaliknya
        for nama_db, cid in db_mapping.items():
            if nama_db in nama_excel or nama_excel in nama_db:
                return cid
        return None

    df = df.copy()
    df['komoditas_id'] = df['nama_komoditas'].apply(find_id)

    # Laporan matching
    matched   = df[df['komoditas_id'].notna()]
    unmatched = df[df['komoditas_id'].isna()]['nama_komoditas'].unique()

    print(f"\n[Match] {df['nama_komoditas'].nunique()} komoditas di Excel")
    print(f"[Match] {matched['nama_komoditas'].nunique()} berhasil dicocokkan")

    if len(unmatched) > 0:
        print(f"[Match] {len(unmatched)} komoditas TIDAK cocok (akan dilewati):")
        for nama in unmatched[:20]:
            print(f"         - {nama}")

    return matched.dropna(subset=['komoditas_id'])


# ═══════════════════════════════════════════════════════════════
# STEP 4 — Simpan ke database
# ═══════════════════════════════════════════════════════════════

def save_rh_to_db(df: pd.DataFrame, engine) -> dict:
    """
    Simpan data RH ke tabel rh_komoditas.
    Menggunakan upsert (INSERT ... ON DUPLICATE KEY UPDATE).
    """
    if df.empty:
        return {'inserted': 0, 'updated': 0, 'error': 0}

    upsert_query = """
        INSERT INTO rh_komoditas (komoditas_id, tanggal, nilai_rh)
        VALUES (:komoditas_id, :tanggal, :nilai_rh)
        ON DUPLICATE KEY UPDATE
            nilai_rh   = VALUES(nilai_rh),
            updated_at = CURRENT_TIMESTAMP
    """

    inserted = 0
    error    = 0

    with engine.begin() as conn:
        for _, row in df.iterrows():
            try:
                conn.execute(text(upsert_query), {
                    'komoditas_id': int(row['komoditas_id']),
                    'tanggal':      row['tanggal'].strftime('%Y-%m-%d'),
                    'nilai_rh':     float(row['nilai_rh']),
                })
                inserted += 1
            except Exception as e:
                print(f"[DB] Error row {row['nama_komoditas']} {row['tanggal']}: {e}")
                error += 1

    return {'inserted': inserted, 'error': error}


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  IMPORT DATA RH DARI EXCEL BPS -> DATABASE")
    print("=" * 60)

    # Cek file Excel
    if not os.path.exists(EXCEL_PATH):
        print(f"\n[ERROR] File tidak ditemukan: {EXCEL_PATH}")
        print(f"        Letakkan file Excel di folder yang sama dengan script ini,")
        print(f"        atau ubah variabel EXCEL_PATH di bagian atas script.")
        sys.exit(1)

    # Koneksi database
    print(f"\n[DB] Menghubungkan ke database...")
    engine = get_engine()
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        print(f"[DB] Koneksi berhasil!")
    except Exception as e:
        print(f"[DB] Gagal koneksi: {e}")
        sys.exit(1)

    # Load mapping komoditas dari DB
    db_mapping = load_komoditas_from_db(engine)

    # Baca RH dari Excel
    df_rh = read_rh_from_excel(EXCEL_PATH, SHEET_NAME)

    if df_rh.empty:
        print("\n[ERROR] Tidak ada data RH yang terbaca dari Excel.")
        sys.exit(1)

    # Match komoditas
    df_matched = match_komoditas(df_rh, db_mapping)

    if df_matched.empty:
        print("\n[ERROR] Tidak ada komoditas yang cocok. "
              "Periksa nama komoditas di Excel vs database.")
        sys.exit(1)

    # Preview data sebelum simpan
    print(f"\n[Preview] Sample data yang akan disimpan:")
    sample = df_matched.groupby('nama_komoditas').first().reset_index()
    for _, row in sample.head(5).iterrows():
        print(f"  ID {int(row['komoditas_id']):2d} | {row['nama_komoditas'][:30]:30s} | "
              f"{row['tanggal']} | RH={row['nilai_rh']:.4f}")

    # Konfirmasi
    total = len(df_matched)
    print(f"\n[Konfirmasi] Akan menyimpan {total} records ke tabel rh_komoditas.")
    confirm = input("Lanjutkan? (y/n): ").strip().lower()
    if confirm != 'y':
        print("Import dibatalkan.")
        sys.exit(0)

    # Simpan ke DB
    print(f"\n[DB] Menyimpan data...")
    result = save_rh_to_db(df_matched, engine)

    print(f"\n{'=' * 60}")
    print(f"  HASIL IMPORT")
    print(f"{'=' * 60}")
    print(f"  Berhasil disimpan : {result['inserted']} records")
    print(f"  Error             : {result['error']} records")
    print(f"{'=' * 60}")
    print(f"\nSelesai! Data RH siap digunakan untuk kalkulasi IHK.")


if __name__ == '__main__':
    main()